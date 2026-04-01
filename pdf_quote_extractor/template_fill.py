from __future__ import annotations

import csv
from copy import copy
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .normalize import parse_currency_value, parse_number_value


DEFAULT_TEMPLATE_SHEET = "QuoteExportResults"
DEFAULT_TEMPLATE_HEADER_ROW = 4
DEFAULT_TEMPLATE_DATA_START_ROW = 5

DEFAULT_BUSINESS_UNIT = "Spain"
DEFAULT_CURRENCY = "EUR"
DEFAULT_LOCATION = "EXN Spain : ES Sales Stock"

CLIENT_MANAGED_HEADERS = {
    "ExternalId",
    "Title",
    "Reseller",
    "ResellerContact",
    "EndUser",
    "VendorSpecialPriceApproval",
    "VendorSpecialPriceApproval (Line)",
    "SalesExchangeRate",
}

CANONICAL_TEMPLATE_HEADERS = [
    "ExternalId",
    "Title",
    "Currency",
    "Date",
    "Reseller",
    "ResellerContact",
    "Expires",
    "ExpectedClose",
    "EndUser",
    "BusinessUnit",
    "Item",
    "Quantity",
    "Salesprice",
    "Salesdiscount",
    "Purchaseprice",
    "PurchaseDiscount",
    "Location",
    "ContractStart",
    "ContractEnd",
    "Serial#Supported",
    "Rebate",
    "Opportunity",
    "Memo (Line)",
    "Quote ID (Line)",
    "VendorSpecialPriceApproval",
    "VendorSpecialPriceApproval (Line)",
    "SalesCurrency",
    "SalesExchangeRate",
]

REQUIRED_TEMPLATE_HEADERS = [
    "Date",
    "Expires",
    "ExpectedClose",
    "Item",
    "Quantity",
    "Salesprice",
    "Salesdiscount",
    "Purchaseprice",
    "PurchaseDiscount",
    "ContractStart",
    "ContractEnd",
    "Serial#Supported",
    "Rebate",
    "Opportunity",
    "Memo (Line)",
    "Quote ID (Line)",
    "SalesCurrency",
    "SalesExchangeRate",
]

HEADER_NUMBER_FORMATS: dict[str, str] = {
    "Quantity": "0",
    "Salesprice": "0.00",
    "Salesdiscount": "0.00%",
    "Purchaseprice": "0.00",
    "PurchaseDiscount": "0.00%",
}

HEADER_ALIASES: dict[str, set[str]] = {
    "ExternalId": {"externalid", "external id", "internalid", "internal id"},
    "Title": {"title"},
    "Currency": {"currency"},
    "Date": {"date"},
    "Reseller": {"reseller"},
    "ResellerContact": {"resellercontact", "reseller contact"},
    "Expires": {"expires"},
    "ExpectedClose": {"expectedclose"},
    "EndUser": {"enduser", "end user"},
    "BusinessUnit": {"businessunit", "business unit"},
    "Item": {"item"},
    "Quantity": {"quantity", "qty"},
    "Salesprice": {"salesprice", "sales_price", "sales price"},
    "Salesdiscount": {"salesdiscount", "sales_discount", "sales discount"},
    "Purchaseprice": {"purchaseprice", "purchase_price", "purchase price"},
    "PurchaseDiscount": {"purchasediscount", "purchase_discount", "purchase discount"},
    "Location": {"location"},
    "ContractStart": {"contractstart", "contract_start", "contract start"},
    "ContractEnd": {"contractend", "contract_end", "contract end"},
    "Serial#Supported": {"serialsupported", "serial", "serialnumbersupported"},
    "Rebate": {"rebate"},
    "Opportunity": {"opportunity"},
    "Memo (Line)": {"memoline", "memo"},
    "Quote ID (Line)": {"quoteidline", "quoteid", "quote id"},
    "VendorSpecialPriceApproval": {
        "vendorspecialpriceapproval",
        "vendor special price approval",
    },
    "VendorSpecialPriceApproval (Line)": {
        "vendorspecialpriceapprovalline",
        "vendorspecialpriceapprovalline",
        "vendor special price approval line",
    },
    "SalesCurrency": {"salescurrency", "sales currency"},
    "SalesExchangeRate": {"salesexchangerate", "sales exchange rate"},
}

NORMALIZED_HEADER_ALIASES: dict[str, set[str]] = {
    target: {re.sub(r"[^a-z0-9]+", "", alias.strip().lower()) for alias in aliases}
    for target, aliases in HEADER_ALIASES.items()
}


def _clean_single_line(text: str | None) -> str | None:
    if text is None:
        return None
    cleaned = " ".join(str(text).replace("\n", " ").split())
    return cleaned or None


def _clean_sku(text: str | None) -> str | None:
    single_line = _clean_single_line(text)
    if single_line is None:
        return None
    cleaned = re.sub(r"\s+", "", single_line)
    return cleaned or None


def _format_date_for_template(raw: str | None) -> str | None:
    if not raw:
        return None
    text = str(raw).strip()
    if not text:
        return None
    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return text


def _normalize_header(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", text.strip().lower())


def _parse_quantity(raw: str | None) -> float | int | None:
    if not raw:
        return None
    value = parse_number_value(raw, allow_thousands=True)
    if value is None:
        return None
    if value.is_integer():
        return int(value)
    return value


def _parse_discount_fraction(
    discount_pct_value: float | None,
    discount_pct_raw: str | None,
) -> float | None:
    if discount_pct_value is not None:
        return round(float(discount_pct_value) / 100.0, 6)
    if discount_pct_raw:
        parsed = parse_number_value(discount_pct_raw, allow_thousands=False)
        if parsed is not None:
            return round(parsed / 100.0, 6)
    return None


def _parse_price_value(value: float | None, raw: str | None) -> float | None:
    if value is not None:
        return float(value)
    return parse_currency_value(raw)


def _derive_cost_basis(
    purchase_price_display: float | None,
    purchase_discount_fraction: float | None,
    net_unit_price: float | None,
) -> float | None:
    if net_unit_price is not None:
        return net_unit_price
    if purchase_price_display is None:
        return None
    if purchase_discount_fraction is None:
        return purchase_price_display
    return round(purchase_price_display * (1.0 - purchase_discount_fraction), 6)


def _parse_sales_discount(
    sales_price: float | None,
    purchase_cost_basis: float | None,
    euro_rate: float,
    margin_percent: float,
) -> float | None:
    if sales_price is None or purchase_cost_basis is None:
        return None
    if sales_price == 0:
        return None
    if euro_rate <= 0:
        return None
    margin_divisor = 1.0 - (margin_percent / 100.0)
    value = 1.0 - (((purchase_cost_basis / euro_rate) / margin_divisor) / sales_price)
    return round(value, 6)


def _build_template_rows(
    files_payload: list[dict[str, Any]],
    euro_rate: float,
    margin_percent: float,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for file_payload in files_payload:
        summary = file_payload.get("business_summary", {})
        line_items = file_payload.get("line_items_parsed", [])

        quote_id = summary.get("quote_number")
        expiration_date = _format_date_for_template(summary.get("expiration_date"))
        for item in line_items:
            discount_fraction = _parse_discount_fraction(
                item.get("discount_pct_value"),
                item.get("discount_pct_raw"),
            )
            sales_price = _parse_price_value(
                item.get("list_unit_price_value"),
                item.get("list_unit_price_raw"),
            )
            purchase_price = sales_price
            if purchase_price is None:
                purchase_price = _parse_price_value(
                    item.get("net_unit_price_value"),
                    item.get("net_unit_price_raw"),
                )
            purchase_cost_basis = _derive_cost_basis(
                purchase_price_display=purchase_price,
                purchase_discount_fraction=discount_fraction,
                net_unit_price=_parse_price_value(
                    item.get("net_unit_price_value"),
                    item.get("net_unit_price_raw"),
                ),
            )
            sales_discount = _parse_sales_discount(
                sales_price=sales_price,
                purchase_cost_basis=purchase_cost_basis,
                euro_rate=euro_rate,
                margin_percent=margin_percent,
            )

            row = {
                "ExternalId": None,
                "Title": None,
                "Currency": DEFAULT_CURRENCY,
                "Date": None,
                "Reseller": None,
                "ResellerContact": None,
                "Expires": expiration_date,
                "ExpectedClose": expiration_date,
                "EndUser": None,
                "BusinessUnit": DEFAULT_BUSINESS_UNIT,
                "Item": _clean_sku(item.get("sku")),
                "Quantity": _parse_quantity(item.get("units_qty")),
                "Salesprice": sales_price,
                "Salesdiscount": sales_discount,
                "Purchaseprice": purchase_price,
                "PurchaseDiscount": discount_fraction,
                "Location": DEFAULT_LOCATION,
                "ContractStart": _format_date_for_template(item.get("term_start")),
                "ContractEnd": _format_date_for_template(item.get("term_end")),
                "Serial#Supported": None,
                "Rebate": None,
                "Opportunity": None,
                "Memo (Line)": None,
                "Quote ID (Line)": quote_id,
                "VendorSpecialPriceApproval": None,
                "VendorSpecialPriceApproval (Line)": None,
                "SalesCurrency": DEFAULT_CURRENCY,
                "SalesExchangeRate": None,
            }
            rows.append(row)
    return rows


def _format_decimal_for_csv(value: float, decimals: int = 2) -> str:
    return f"{value:.{decimals}f}".replace(".", ",")


def _format_percentage_for_csv(value: float) -> str:
    return f"{value * 100:.2f}%".replace(".", ",")


def _format_csv_cell(header: str, value: Any) -> str:
    if value in (None, ""):
        return ""
    if header == "Quantity":
        if isinstance(value, float) and not value.is_integer():
            return _format_decimal_for_csv(value)
        if isinstance(value, int | float):
            return str(int(value))
    if header in {"Salesprice", "Purchaseprice", "SalesExchangeRate"} and isinstance(value, int | float):
        return _format_decimal_for_csv(float(value))
    if header in {"Salesdiscount", "PurchaseDiscount"} and isinstance(value, int | float):
        return _format_percentage_for_csv(float(value))
    return str(value)


def _match_headers_in_row(ws, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col_idx).value
        if isinstance(value, str):
            normalized = _normalize_header(value)
            for target_header, aliases in NORMALIZED_HEADER_ALIASES.items():
                if normalized in aliases and target_header not in headers:
                    headers[target_header] = col_idx
                    break
    return headers


def _copy_column_template(ws, source_col_idx: int, target_col_idx: int) -> None:
    source_letter = get_column_letter(source_col_idx)
    target_letter = get_column_letter(target_col_idx)
    source_dimension = ws.column_dimensions[source_letter]
    target_dimension = ws.column_dimensions[target_letter]
    target_dimension.width = source_dimension.width
    target_dimension.hidden = source_dimension.hidden
    target_dimension.bestFit = source_dimension.bestFit
    target_dimension.outlineLevel = source_dimension.outlineLevel
    target_dimension.collapsed = source_dimension.collapsed

    for row_idx in range(1, ws.max_row + 1):
        source_cell = ws.cell(row_idx, source_col_idx)
        target_cell = ws.cell(row_idx, target_col_idx)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)


def _copy_row_template(ws, source_row_idx: int, target_row_idx: int) -> None:
    source_dimension = ws.row_dimensions[source_row_idx]
    target_dimension = ws.row_dimensions[target_row_idx]
    target_dimension.height = source_dimension.height
    target_dimension.hidden = source_dimension.hidden
    target_dimension.outlineLevel = source_dimension.outlineLevel
    target_dimension.collapsed = source_dimension.collapsed

    for col_idx in range(1, ws.max_column + 1):
        source_cell = ws.cell(source_row_idx, col_idx)
        target_cell = ws.cell(target_row_idx, col_idx)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)


def _canonicalize_template_layout(ws, header_row: int, headers: dict[str, int]) -> dict[str, int]:
    if "Quote ID (Line)" in headers:
        expected_col = headers["Quote ID (Line)"] + 1
        for header in (
            "VendorSpecialPriceApproval",
            "VendorSpecialPriceApproval (Line)",
            "SalesCurrency",
            "SalesExchangeRate",
        ):
            actual_col = headers.get(header)
            if actual_col is None:
                ws.insert_cols(expected_col, amount=1)
                source_col_idx = min(expected_col + 1, ws.max_column)
                _copy_column_template(ws, source_col_idx, expected_col)
                ws.cell(header_row, expected_col).value = header
                headers = _match_headers_in_row(ws, header_row)
                actual_col = headers.get(header)
            if actual_col != expected_col:
                break
            expected_col += 1

    headers = _match_headers_in_row(ws, header_row)
    for header, col_idx in headers.items():
        ws.cell(header_row, col_idx).value = header
    return _match_headers_in_row(ws, header_row)


def _ensure_row_capacity(ws, data_start_row: int, required_rows: int) -> int:
    capacity = max(0, ws.max_row - data_start_row + 1)
    if required_rows <= capacity:
        return capacity

    rows_to_add = required_rows - capacity
    source_row_idx = max(data_start_row, ws.max_row)
    original_max_row = ws.max_row
    ws.insert_rows(original_max_row + 1, amount=rows_to_add)
    for offset in range(rows_to_add):
        _copy_row_template(ws, source_row_idx, original_max_row + offset + 1)
    return max(0, ws.max_row - data_start_row + 1)


def _resolve_header_row_and_columns(ws, header_row: int | None) -> tuple[int, dict[str, int]]:
    if header_row is not None:
        headers = _match_headers_in_row(ws, header_row)
        if all(name in headers for name in REQUIRED_TEMPLATE_HEADERS):
            return header_row, _canonicalize_template_layout(ws, header_row, headers)

    max_scan_row = min(ws.max_row, 50)
    best_row = None
    best_headers: dict[str, int] = {}
    for row_idx in range(1, max_scan_row + 1):
        headers = _match_headers_in_row(ws, row_idx)
        if len(headers) > len(best_headers):
            best_headers = headers
            best_row = row_idx
        if all(name in headers for name in REQUIRED_TEMPLATE_HEADERS):
            return row_idx, _canonicalize_template_layout(ws, row_idx, headers)

    missing = [name for name in REQUIRED_TEMPLATE_HEADERS if name not in best_headers]
    missing_str = ", ".join(missing)
    if best_row is not None:
        raise ValueError(
            f"Template headers not fully found. Best match row={best_row}, missing: {missing_str}"
        )
    raise ValueError(f"Template headers not found. Missing: {missing_str}")


def fill_quote_template(
    template_path: Path,
    template_output_path: Path,
    files_payload: list[dict[str, Any]],
    euro_rate: float | None,
    margin_percent: float | None,
    sheet_name: str = DEFAULT_TEMPLATE_SHEET,
    header_row: int | None = None,
    data_start_row: int | None = None,
) -> dict[str, Any]:
    if euro_rate is None or euro_rate <= 0:
        raise ValueError("euro_rate must be provided and greater than 0.")
    if margin_percent is None:
        raise ValueError("margin_percent must be provided.")

    wb = load_workbook(template_path)
    resolved_sheet_name = sheet_name
    if sheet_name not in wb.sheetnames:
        if len(wb.sheetnames) == 1:
            resolved_sheet_name = wb.sheetnames[0]
        else:
            raise ValueError(f"Template sheet not found: {sheet_name}")
    ws = wb[resolved_sheet_name]
    resolved_header_row, header_columns = _resolve_header_row_and_columns(ws, header_row)

    if data_start_row is None:
        resolved_data_start_row = resolved_header_row + 1
    else:
        resolved_data_start_row = data_start_row
    if resolved_data_start_row <= resolved_header_row:
        raise ValueError(
            f"data_start_row ({resolved_data_start_row}) must be greater than header_row ({resolved_header_row})."
        )

    rows_to_write = _build_template_rows(
        files_payload=files_payload,
        euro_rate=float(euro_rate),
        margin_percent=float(margin_percent),
    )
    capacity = _ensure_row_capacity(ws, resolved_data_start_row, len(rows_to_write))

    # Clear previous data only in yellow target columns.
    for row_idx in range(resolved_data_start_row, ws.max_row + 1):
        for col_idx in header_columns.values():
            ws.cell(row_idx, col_idx).value = None

    # Fill rows from extracted line items.
    for offset, row_values in enumerate(rows_to_write):
        row_idx = resolved_data_start_row + offset
        for header, col_idx in header_columns.items():
            cell = ws.cell(row_idx, col_idx)
            value = row_values.get(header)
            cell.value = value
            number_format = HEADER_NUMBER_FORMATS.get(header)
            if number_format and value not in (None, ""):
                cell.number_format = number_format

    wb.save(template_output_path)
    return {
        "template_path": str(template_path),
        "template_output_path": str(template_output_path),
        "sheet_name": resolved_sheet_name,
        "rows_written": len(rows_to_write),
        "capacity": capacity,
        "euro_rate": float(euro_rate),
        "margin_percent": float(margin_percent),
        "header_row": resolved_header_row,
        "data_start_row": resolved_data_start_row,
    }


def write_quote_csv(
    output_path: Path,
    files_payload: list[dict[str, Any]],
    euro_rate: float | None,
    margin_percent: float | None,
    delimiter: str = ";",
) -> dict[str, Any]:
    if euro_rate is None or euro_rate <= 0:
        raise ValueError("euro_rate must be provided and greater than 0.")
    if margin_percent is None:
        raise ValueError("margin_percent must be provided.")

    rows_to_write = _build_template_rows(
        files_payload=files_payload,
        euro_rate=float(euro_rate),
        margin_percent=float(margin_percent),
    )

    with output_path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.writer(csv_file, delimiter=delimiter, lineterminator="\r\n")
        writer.writerow(CANONICAL_TEMPLATE_HEADERS)
        for row in rows_to_write:
            writer.writerow([_format_csv_cell(header, row.get(header)) for header in CANONICAL_TEMPLATE_HEADERS])

    return {
        "template_output_path": str(output_path),
        "rows_written": len(rows_to_write),
        "format": "csv",
        "delimiter": delimiter,
    }
