import csv
import io
from pathlib import Path
import shutil

from openpyxl import load_workbook

from pdf_quote_extractor.pipeline import run_pipeline
from pdf_quote_extractor.template_fill import _build_template_rows


EXPECTED_TEMPLATE_HEADERS = [
    "ExternalId",
    "Title",
    "Currency",
    "Date",
    "Reseller",
    "ResellerContact",
    "Expires",
    "ExpectedClose",
    "EndUser",
    "BusinessUnit",
    "Item",
    "Quantity",
    "Salesprice",
    "Salesdiscount",
    "Purchaseprice",
    "PurchaseDiscount",
    "Location",
    "ContractStart",
    "ContractEnd",
    "Serial#Supported",
    "Rebate",
    "Opportunity",
    "Memo (Line)",
    "Quote ID (Line)",
    "VendorSpecialPriceApproval",
    "VendorSpecialPriceApproval (Line)",
    "SalesCurrency",
    "SalesExchangeRate",
]


def test_template_fill_for_quote_fixture(tmp_path):
    project_root = Path(__file__).resolve().parents[1]
    fixture_pdf = project_root / "tests" / "fixtures" / "Q-220053-20251224-0752  (1).pdf"
    template_file = project_root / "tests" / "fixtures" / "Example with calculations.xlsx"
    config_path = project_root / "config.yaml"

    output_xlsx = tmp_path / "quote_output.xlsx"
    output_json = tmp_path / "quote_output.json"
    filled_template = tmp_path / "quote_template_filled.xlsx"

    exit_code, payload = run_pipeline(
        input_path=fixture_pdf,
        output_path=output_xlsx,
        json_output_path=output_json,
        config_path=config_path,
        ocr_mode="auto",
        strict=True,
        include_char_layer=False,
        include_tables=True,
        tesseract_cmd=None,
        poppler_path=None,
        template_path=template_file,
        template_output_path=filled_template,
        euro_rate=1.17,
        margin_percent=10.0,
    )

    assert exit_code == 0
    assert payload["template_output"]["rows_written"] == 4
    assert filled_template.exists()

    wb = load_workbook(filled_template, data_only=False)
    ws = wb["QuoteExportResults"]
    headers = [ws.cell(4, col).value for col in range(1, len(EXPECTED_TEMPLATE_HEADERS) + 1)]
    assert headers == EXPECTED_TEMPLATE_HEADERS

    # Row 5: first parsed line item.
    assert ws["C5"].value == "EUR"
    assert ws["D5"].value in (None, "")
    assert ws["G5"].value == "30/12/2025"
    assert ws["H5"].value == "30/12/2025"
    assert ws["J5"].value == "Spain"
    assert ws["K5"].value == "NK-EGRESS-DIP"
    assert ws["L5"].value == 10000
    assert ws["M5"].value == 60
    assert ws["N5"].value == 0.843305
    assert ws["O5"].value == 60
    assert ws["P5"].value == 0.835
    assert ws["N5"].number_format == "0.00%"
    assert ws["P5"].number_format == "0.00%"
    assert ws["Q5"].value == "EXN Spain : ES Sales Stock"
    assert ws["R5"].value == "31/12/2025"
    assert ws["S5"].value == "30/12/2028"
    assert ws["V5"].value in (None, "")
    assert ws["W5"].value in (None, "")
    assert ws["X5"].value == "Q-220053-2"
    assert ws["Y5"].value in (None, "")
    assert ws["Z5"].value in (None, "")
    assert ws["AA5"].value == "EUR"
    assert ws["AB5"].value in (None, "")

    # Row 6: second parsed line item.
    assert ws["L6"].value == 1
    assert ws["M6"].value == 48000
    assert ws["N6"].value == 0.498575
    assert ws["O6"].value == 48000
    assert ws["P6"].value == 0.472
    assert ws["N6"].number_format == "0.00%"
    assert ws["P6"].number_format == "0.00%"
    assert ws["K6"].value == "NK-SSLI"
    assert ws["V6"].value in (None, "")
    assert ws["W6"].value in (None, "")
    assert ws["Y6"].value in (None, "")
    assert ws["Z6"].value in (None, "")
    assert ws["AA6"].value == "EUR"


def test_template_only_mode_skips_audit_workbook(tmp_path):
    project_root = Path(__file__).resolve().parents[1]
    fixture_pdf = project_root / "tests" / "fixtures" / "Q-220053-20251224-0752  (1).pdf"
    template_file = project_root / "tests" / "fixtures" / "Example with calculations.xlsx"
    config_path = project_root / "config.yaml"

    output_xlsx = tmp_path / "quote_output.xlsx"
    output_json = tmp_path / "quote_output.json"
    filled_template = tmp_path / "quote_template_filled.xlsx"

    exit_code, payload = run_pipeline(
        input_path=fixture_pdf,
        output_path=output_xlsx,
        json_output_path=output_json,
        config_path=config_path,
        ocr_mode="auto",
        strict=True,
        include_char_layer=False,
        include_tables=True,
        tesseract_cmd=None,
        poppler_path=None,
        template_path=template_file,
        template_output_path=filled_template,
        euro_rate=1.17,
        margin_percent=10.0,
        write_audit_workbook=False,
    )

    assert exit_code == 0
    assert payload["file_count"] == 1
    assert filled_template.exists()
    assert not output_xlsx.exists()


def test_csv_export_writes_client_ready_rows(tmp_path):
    project_root = Path(__file__).resolve().parents[1]
    fixture_pdf = project_root / "tests" / "fixtures" / "Q-220053-20251224-0752  (1).pdf"
    config_path = project_root / "config.yaml"

    output_xlsx = tmp_path / "quote_output.xlsx"
    output_json = tmp_path / "quote_output.json"
    csv_output = tmp_path / "quote_template_filled.csv"

    exit_code, payload = run_pipeline(
        input_path=fixture_pdf,
        output_path=output_xlsx,
        json_output_path=output_json,
        config_path=config_path,
        ocr_mode="off",
        strict=True,
        include_char_layer=False,
        include_tables=True,
        tesseract_cmd=None,
        poppler_path=None,
        template_path=None,
        template_output_path=csv_output,
        euro_rate=1.17,
        margin_percent=10.0,
        write_audit_workbook=False,
    )

    assert exit_code == 0
    assert payload["template_output"]["rows_written"] == 4
    assert payload["template_output"]["format"] == "csv"
    assert csv_output.exists()

    rows = list(csv.reader(io.StringIO(csv_output.read_text(encoding="utf-8-sig")), delimiter=";"))
    assert rows[0] == EXPECTED_TEMPLATE_HEADERS
    assert len(rows) == 5
    assert rows[1][3] == ""
    assert rows[1][10] == "NK-EGRESS-DIP"
    assert rows[1][11] == "10000"
    assert rows[1][12] == "60,00"
    assert rows[1][13] == "84,33%"
    assert rows[1][14] == "60,00"
    assert rows[1][15] == "83,50%"
    assert rows[1][23] == "Q-220053-2"
    assert rows[1][26] == "EUR"


def test_template_fill_with_multiple_pdfs(tmp_path):
    project_root = Path(__file__).resolve().parents[1]
    fixture_pdf = project_root / "tests" / "fixtures" / "Q-220053-20251224-0752  (1).pdf"
    template_file = project_root / "tests" / "fixtures" / "Example with calculations.xlsx"
    config_path = project_root / "config.yaml"

    batch_dir = tmp_path / "batch"
    batch_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(fixture_pdf, batch_dir / "quote_1.pdf")
    shutil.copy2(fixture_pdf, batch_dir / "quote_2.pdf")

    output_xlsx = tmp_path / "quote_output.xlsx"
    output_json = tmp_path / "quote_output.json"
    filled_template = tmp_path / "quote_template_filled.xlsx"

    exit_code, payload = run_pipeline(
        input_path=batch_dir,
        output_path=output_xlsx,
        json_output_path=output_json,
        config_path=config_path,
        ocr_mode="off",
        strict=True,
        include_char_layer=False,
        include_tables=True,
        tesseract_cmd=None,
        poppler_path=None,
        template_path=template_file,
        template_output_path=filled_template,
        euro_rate=1.17,
        margin_percent=10.0,
        write_audit_workbook=False,
    )

    assert exit_code == 0
    assert payload["file_count"] == 2
    assert payload["template_output"]["rows_written"] == 8
    assert filled_template.exists()

    wb = load_workbook(filled_template, data_only=False)
    ws = wb["QuoteExportResults"]
    assert ws["K5"].value is not None
    assert ws["K12"].value is not None


def test_template_fill_with_multiple_input_paths(tmp_path):
    project_root = Path(__file__).resolve().parents[1]
    fixture_pdf = project_root / "tests" / "fixtures" / "Q-220053-20251224-0752  (1).pdf"
    template_file = project_root / "tests" / "fixtures" / "Example with calculations.xlsx"
    config_path = project_root / "config.yaml"

    pdf1 = tmp_path / "input_1.pdf"
    pdf2 = tmp_path / "input_2.pdf"
    shutil.copy2(fixture_pdf, pdf1)
    shutil.copy2(fixture_pdf, pdf2)

    output_xlsx = tmp_path / "quote_output.xlsx"
    output_json = tmp_path / "quote_output.json"
    filled_template = tmp_path / "quote_template_filled.xlsx"

    exit_code, payload = run_pipeline(
        input_path=[pdf1, pdf2],
        output_path=output_xlsx,
        json_output_path=output_json,
        config_path=config_path,
        ocr_mode="off",
        strict=True,
        include_char_layer=False,
        include_tables=True,
        tesseract_cmd=None,
        poppler_path=None,
        template_path=template_file,
        template_output_path=filled_template,
        euro_rate=1.17,
        margin_percent=10.0,
        write_audit_workbook=False,
    )

    assert exit_code == 0
    assert payload["file_count"] == 2
    assert payload["template_output"]["rows_written"] == 8


def test_template_autodetects_header_row_and_data_start(tmp_path):
    project_root = Path(__file__).resolve().parents[1]
    fixture_pdf = project_root / "tests" / "fixtures" / "Q-220053-20251224-0752  (1).pdf"
    source_template = project_root / "tests" / "fixtures" / "Example with calculations.xlsx"
    config_path = project_root / "config.yaml"

    template_file = tmp_path / "template_headers_row2.xlsx"
    shutil.copy2(source_template, template_file)

    wb = load_workbook(template_file, data_only=False)
    ws = wb["QuoteExportResults"]
    for col in range(1, len(EXPECTED_TEMPLATE_HEADERS) + 1):
        ws.cell(2, col).value = ws.cell(4, col).value
        ws.cell(4, col).value = None
    wb.save(template_file)

    output_xlsx = tmp_path / "quote_output.xlsx"
    output_json = tmp_path / "quote_output.json"
    filled_template = tmp_path / "quote_template_filled.xlsx"

    exit_code, payload = run_pipeline(
        input_path=fixture_pdf,
        output_path=output_xlsx,
        json_output_path=output_json,
        config_path=config_path,
        ocr_mode="off",
        strict=True,
        include_char_layer=False,
        include_tables=True,
        tesseract_cmd=None,
        poppler_path=None,
        template_path=template_file,
        template_output_path=filled_template,
        euro_rate=1.17,
        margin_percent=10.0,
        write_audit_workbook=False,
    )

    assert exit_code == 0
    assert payload["template_output"]["header_row"] == 2
    assert payload["template_output"]["data_start_row"] == 3

    wb_out = load_workbook(filled_template, data_only=False)
    ws_out = wb_out["QuoteExportResults"]
    headers = [ws_out.cell(2, col).value for col in range(1, len(EXPECTED_TEMPLATE_HEADERS) + 1)]
    assert headers == EXPECTED_TEMPLATE_HEADERS
    assert ws_out["K3"].value is not None
    assert ws_out["M3"].value == 60
    assert ws_out["N3"].value == 0.843305
    assert ws_out["N3"].number_format == "0.00%"
    assert ws_out["P3"].number_format == "0.00%"
    assert ws_out["D3"].value in (None, "")
    assert ws_out["V3"].value in (None, "")
    assert ws_out["X3"].value == "Q-220053-2"
    assert ws_out["Y3"].value in (None, "")
    assert ws_out["Z3"].value in (None, "")
    assert ws_out["AA3"].value == "EUR"


def test_template_fill_falls_back_to_single_sheet_when_default_name_is_missing(tmp_path):
    project_root = Path(__file__).resolve().parents[1]
    fixture_pdf = project_root / "tests" / "fixtures" / "Q-220053-20251224-0752  (1).pdf"
    source_template = project_root / "tests" / "fixtures" / "Example with calculations.xlsx"
    config_path = project_root / "config.yaml"

    template_file = tmp_path / "template_single_sheet_renamed.xlsx"
    shutil.copy2(source_template, template_file)

    wb = load_workbook(template_file, data_only=False)
    ws = wb["QuoteExportResults"]
    ws.title = "Client Template"
    wb.save(template_file)

    output_xlsx = tmp_path / "quote_output.xlsx"
    output_json = tmp_path / "quote_output.json"
    filled_template = tmp_path / "quote_template_filled.xlsx"

    exit_code, payload = run_pipeline(
        input_path=fixture_pdf,
        output_path=output_xlsx,
        json_output_path=output_json,
        config_path=config_path,
        ocr_mode="off",
        strict=True,
        include_char_layer=False,
        include_tables=True,
        tesseract_cmd=None,
        poppler_path=None,
        template_path=template_file,
        template_output_path=filled_template,
        euro_rate=1.17,
        margin_percent=10.0,
        write_audit_workbook=False,
    )

    assert exit_code == 0
    assert payload["template_output"]["sheet_name"] == "Client Template"

    wb_out = load_workbook(filled_template, data_only=False)
    assert "Client Template" in wb_out.sheetnames


def test_template_fill_upgrades_legacy_salescurrency_layout(tmp_path):
    project_root = Path(__file__).resolve().parents[1]
    fixture_pdf = project_root / "tests" / "fixtures" / "Q-220053-20251224-0752  (1).pdf"
    source_template = project_root / "tests" / "fixtures" / "Example with calculations.xlsx"
    config_path = project_root / "config.yaml"

    template_file = tmp_path / "template_legacy_layout.xlsx"
    shutil.copy2(source_template, template_file)

    wb = load_workbook(template_file, data_only=False)
    ws = wb["QuoteExportResults"]
    ws.delete_cols(25, 2)
    wb.save(template_file)

    output_xlsx = tmp_path / "quote_output.xlsx"
    output_json = tmp_path / "quote_output.json"
    filled_template = tmp_path / "quote_template_filled.xlsx"

    exit_code, payload = run_pipeline(
        input_path=fixture_pdf,
        output_path=output_xlsx,
        json_output_path=output_json,
        config_path=config_path,
        ocr_mode="off",
        strict=True,
        include_char_layer=False,
        include_tables=True,
        tesseract_cmd=None,
        poppler_path=None,
        template_path=template_file,
        template_output_path=filled_template,
        euro_rate=1.17,
        margin_percent=10.0,
        write_audit_workbook=False,
    )

    assert exit_code == 0
    assert payload["template_output"]["rows_written"] == 4

    wb_out = load_workbook(filled_template, data_only=False)
    ws_out = wb_out["QuoteExportResults"]
    headers = [ws_out.cell(4, col).value for col in range(1, len(EXPECTED_TEMPLATE_HEADERS) + 1)]
    assert headers == EXPECTED_TEMPLATE_HEADERS
    assert ws_out["X5"].value == "Q-220053-2"
    assert ws_out["Y5"].value in (None, "")
    assert ws_out["Z5"].value in (None, "")
    assert ws_out["AA5"].value == "EUR"
    assert ws_out["AB5"].value in (None, "")


def test_template_rows_include_included_zero_value_items():
    files_payload = [
        {
            "metadata": {"creation_date": "D:20260101000000Z"},
            "business_summary": {"quote_number": "Q-TEST", "expiration_date": "01/31/2026"},
            "line_items_parsed": [
                {
                    "sku": "SKU-A",
                    "units_qty": "1",
                    "term_start": "01/01/2026",
                    "term_end": "01/31/2026",
                    "list_unit_price_value": 100.0,
                    "discount_pct_value": 0.0,
                    "net_unit_price_value": 100.0,
                    "net_total_value": 100.0,
                },
                {
                    "sku": "SKU-B",
                    "units_qty": "1",
                    "term_start": "01/01/2026",
                    "term_end": "01/31/2026",
                    "list_unit_price_raw": "USD 0.00",
                    "discount_pct_raw": "0.00",
                    "net_unit_price_raw": "Included",
                    "net_total_raw": "USD 0.00",
                },
            ],
        }
    ]

    rows = _build_template_rows(files_payload=files_payload, euro_rate=1.17, margin_percent=20.0)

    assert len(rows) == 2
    assert rows[0]["Item"] == "SKU-A"
    assert rows[1]["Item"] == "SKU-B"
    assert rows[1]["Salesprice"] == 0.0
    assert rows[1]["Salesdiscount"] is None
    assert rows[0]["Purchaseprice"] == 100.0
    assert rows[1]["Purchaseprice"] == 0.0
    assert rows[0]["Date"] is None
    assert rows[0]["Expires"] == "31/01/2026"
    assert rows[0]["ExpectedClose"] == "31/01/2026"
    assert rows[0]["ContractStart"] == "01/01/2026"
    assert rows[0]["ContractEnd"] == "31/01/2026"
    assert rows[0]["BusinessUnit"] == "Spain"
    assert rows[0]["Currency"] == "EUR"
    assert rows[0]["Location"] == "EXN Spain : ES Sales Stock"
    assert rows[0]["SalesCurrency"] == "EUR"
    assert rows[0]["Opportunity"] is None
    assert rows[0]["Quote ID (Line)"] == "Q-TEST"


def test_template_rows_keep_margin_based_on_net_cost_when_purchaseprice_displays_list():
    files_payload = [
        {
            "metadata": {"creation_date": "D:20260101000000Z"},
            "business_summary": {"quote_number": "Q-MARGIN", "expiration_date": "01/31/2026"},
            "line_items_parsed": [
                {
                    "sku": "SKU-MARGIN",
                    "units_qty": "1",
                    "term_start": "01/01/2026",
                    "term_end": "01/31/2026",
                    "list_unit_price_value": 100.0,
                    "discount_pct_value": 10.0,
                    "net_unit_price_value": 90.0,
                }
            ],
        }
    ]

    rows = _build_template_rows(files_payload=files_payload, euro_rate=1.0, margin_percent=0.0)

    assert rows[0]["Purchaseprice"] == 100.0
    assert rows[0]["PurchaseDiscount"] == 0.1
    assert rows[0]["Salesdiscount"] == 0.1


def test_template_rows_can_derive_margin_from_list_and_discount_without_net_unit_price():
    files_payload = [
        {
            "metadata": {"creation_date": "D:20260101000000Z"},
            "business_summary": {"quote_number": "Q-DERIVE", "expiration_date": "01/31/2026"},
            "line_items_parsed": [
                {
                    "sku": "SKU-DERIVE",
                    "units_qty": "1",
                    "term_start": "01/01/2026",
                    "term_end": "01/31/2026",
                    "list_unit_price_value": 100.0,
                    "discount_pct_value": 10.0,
                }
            ],
        }
    ]

    rows = _build_template_rows(files_payload=files_payload, euro_rate=1.0, margin_percent=0.0)

    assert rows[0]["Purchaseprice"] == 100.0
    assert rows[0]["PurchaseDiscount"] == 0.1
    assert rows[0]["Salesdiscount"] == 0.1


def test_template_rows_leave_client_managed_fields_blank():
    files_payload = [
        {
            "metadata": {"creation_date": "D:20260101000000Z"},
            "business_summary": {
                "quote_number": "Q-CLIENT",
                "expiration_date": "01/31/2026",
                "external_id": "SHOULD-STAY-BLANK",
                "title": "CLIENT TITLE",
                "reseller": "CLIENT RESELLER",
                "reseller_contact": "CLIENT CONTACT",
                "end_user": "CLIENT END USER",
            },
            "line_items_parsed": [
                {
                    "sku": "SKU-CLIENT",
                    "units_qty": "1",
                    "term_start": "01/01/2026",
                    "term_end": "01/31/2026",
                    "list_unit_price_value": 100.0,
                    "discount_pct_value": 10.0,
                    "net_unit_price_value": 90.0,
                }
            ],
        }
    ]

    rows = _build_template_rows(files_payload=files_payload, euro_rate=1.0, margin_percent=0.0)

    assert rows[0]["ExternalId"] is None
    assert rows[0]["Title"] is None
    assert rows[0]["Reseller"] is None
    assert rows[0]["ResellerContact"] is None
    assert rows[0]["EndUser"] is None
    assert rows[0]["VendorSpecialPriceApproval"] is None
    assert rows[0]["VendorSpecialPriceApproval (Line)"] is None
    assert rows[0]["SalesExchangeRate"] is None
