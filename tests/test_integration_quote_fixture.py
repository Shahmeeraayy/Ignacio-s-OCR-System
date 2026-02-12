from pathlib import Path

from openpyxl import load_workbook

from pdf_quote_extractor.pipeline import run_pipeline


def test_integration_quote_fixture(tmp_path):
    project_root = Path(__file__).resolve().parents[1]
    fixture_pdf = project_root / "tests" / "fixtures" / "Q-220053-20251224-0752  (1).pdf"
    assert fixture_pdf.exists(), f"Fixture not found: {fixture_pdf}"

    output_xlsx = tmp_path / "quote_output.xlsx"
    output_json = tmp_path / "quote_output.json"
    config_path = project_root / "config.yaml"

    exit_code, payload = run_pipeline(
        input_path=fixture_pdf,
        output_path=output_xlsx,
        json_output_path=output_json,
        config_path=config_path,
        ocr_mode="auto",
        strict=True,
        include_char_layer=False,
        include_tables=True,
        tesseract_cmd=None,
        poppler_path=None,
    )

    assert exit_code == 0
    assert payload["file_count"] == 1
    file_payload = payload["files"][0]
    assert file_payload["metadata"]["pages"] == 4
    assert file_payload["metadata"]["creator"] == "Apache FOP Version 2.2"

    links_page_3 = [row for row in file_payload["links"] if row["page"] == 3]
    assert len(links_page_3) == 4

    summary = file_payload["business_summary"]
    assert summary["total_value"] == 617572.80
    assert summary["overall_total_value"] == 617572.80
    assert summary["line_items_total_value"] == 617572.80

    critical_failures = [
        row
        for row in file_payload["validation_report"]
        if row["severity"] == "critical" and row["status"] == "FAIL"
    ]
    assert not critical_failures

    assert output_xlsx.exists()
    assert output_json.exists()

    wb = load_workbook(output_xlsx, read_only=True)
    expected_sheets = {
        "document_metadata",
        "pages",
        "text_lines",
        "text_words",
        "tables_raw",
        "line_items_parsed",
        "links",
        "images",
        "business_summary",
        "validation_report",
    }
    assert expected_sheets.issubset(set(wb.sheetnames))

